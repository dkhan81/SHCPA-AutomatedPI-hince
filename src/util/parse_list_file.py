####################################################################
# read a list file(xlsx, xls, csv) and convert it into python list #
####################################################################

from openpyxl import load_workbook
import csv
import xlrd
import pandas as pd

def parse_xlsx(location):
  load_wb = load_workbook(location, data_only=True)
  load_ws = load_wb.active
  
  return list(load_ws.iter_rows(values_only=True))


def parse_csv(location):
  with open(location, 'r') as csv_file:
    parsed_list = []
    
    for row in csv.reader(csv_file):
      parsed_list.append(row)

  return parsed_list


def convert_html_to_list(location):
  tables = pd.read_html(location)
  return tables[0].values.tolist()